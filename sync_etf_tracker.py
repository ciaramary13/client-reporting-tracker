"""
ETF Client Reporting Tracker - Auto Sync Script (multi-sheet)
--------------------------------------------------------------
Syncs THREE separate BNY exports into the THREE matching tabs of the
master ETF Client Reporting Tracker.
"""

import pandas as pd
import openpyxl
import shutil
import os
import sys
import traceback
from datetime import datetime

# ─── CONFIG ────────────────────────────────────────────────────────────────────

TRACKER_PATH = (
    r"C:\Users\ciara\OneDrive - Fordham University\internship-projects"
    r"\client-reporting-tracker\ETF_Client_Reporting_Tracker_V2_DUMMY.xlsx"
)

EXPORT_PATHS = {
    "execution_by_product": r"C:\Users\ciara\Downloads\UserExecutionDetailsByProduct.xlsx",
    "email_baskets":        r"C:\Users\ciara\Downloads\ClientList.xlsx",
    "email_orders":         r"C:\Users\ciara\Downloads\OrderEventExecutionsList.xlsx",
}

LOG_PATH = r"C:\ETF_Sync\sync_log.txt"
SHOW_NOTIFICATIONS = True
PROTECTED_COLS = ["API Identifier"]

JOBS = [
    {
        "name": "Execution by Product",
        "kind": "monthly_columns",
        "export_path_key": "execution_by_product",
        "export_sheet": "Sheet1",
        "tracker_sheet": "Execution by Product",
        "key_cols": ["Client Name", "Product Type", "Report Name"],
        "merge_key": "Client Name",
        "month_format": "banner",
    },
    {
        "name": "Email Events - Baskets",
        "kind": "monthly_columns",
        "export_path_key": "email_baskets",
        "export_sheet": "Sheet1",
        "tracker_sheet": "Email Events - Baskets",
        "key_cols": ["Client"],
        "merge_key": "Client",
        "month_format": "mon_yy",
    },
    {
        "name": "Email Events - Orders",
        "kind": "daily_rows",
        "export_path_key": "email_orders",
        "export_sheet": "Sheet1",
        "tracker_sheet": "Email Events - Orders",
        "key_cols": ["Execution Date"],
        "merge_key": "Execution Date",
        "total_row_label": "Total",
        "sum_cols": ["Sum of Order Count", "Sum of Message Count", "Sum of SSE Executions"],
        "avg_cols": [
            "Min Exec Duration(mins)", "Max Exec Duration(mins)",
            "Avg Exec Duration(mins)", "StdDev of Execution Duration(mins)",
            "Median Exec Duration(mins)", "Successful executions as % of total",
        ],
    },
]

MONTH_NAMES = {
    "jan", "feb", "mar", "apr", "may", "jun",
    "jul", "aug", "sep", "oct", "nov", "dec",
}

_MONTH_ORDER = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"]
)}


def is_month_col(col_name):
    first_word = str(col_name).strip().lower()[:3]
    return first_word in MONTH_NAMES


def has_year_suffix(col_name):
    parts = str(col_name).strip().split()
    return len(parts) > 1 and parts[-1].isdigit() and len(parts[-1]) == 4


def parse_mon_yy(col_name):
    text = str(col_name).strip()
    if "-" not in text:
        return None
    month_part, year_part = text.split("-", 1)
    month_part = month_part.strip()
    year_part = year_part.strip()
    if month_part.lower()[:3] not in MONTH_NAMES or not year_part.isdigit():
        return None
    year_full = 2000 + int(year_part) if len(year_part) == 2 else int(year_part)
    return month_part[:3].title(), year_full


def month_sort_key(col_name, month_format="banner"):
    if month_format == "mon_yy":
        parsed = parse_mon_yy(col_name)
        if parsed:
            abbr, year = parsed
            return (-year, -_MONTH_ORDER.get(abbr.lower(), -1))
        return (0, 0)

    text = str(col_name).strip()
    abbr = text[:3].lower()
    month_num = _MONTH_ORDER.get(abbr, -1)
    parts = text.split()
    year = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
    return (-year, -month_num)


def infer_export_month_years(month_cols, as_of=None):
    if as_of is None:
        as_of = datetime.now()

    year_map = {}
    year, month = as_of.year, as_of.month
    y, m = year, month
    for col in month_cols:
        abbr = str(col).strip().lower()[:3]
        target_month = _MONTH_ORDER.get(abbr)
        if target_month is None:
            continue
        for _ in range(13):
            if m - 1 == target_month:
                year_map[col] = y
                m -= 1
                if m == 0:
                    m = 12
                    y -= 1
                break
            m -= 1
            if m == 0:
                m = 12
                y -= 1
    return year_map

# ─── NOTIFICATIONS / LOGGING ──────────────────────────────────────────────────

def log_line(message):
    os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {message}\n")


def notify(title, message):
    if not SHOW_NOTIFICATIONS:
        return
    try:
        from win10toast_persist import ToastNotifier
        toaster = ToastNotifier()
        toaster.show_toast(title, message, duration=10, threaded=True)
    except ImportError:
        pass
    except Exception:
        pass

# ─── SHARED HELPERS ────────────────────────────────────────────────────────────

def backup_tracker(tracker_path):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = os.path.join(os.path.dirname(tracker_path), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    backup_path = os.path.join(backup_dir, f"ETF_Tracker_backup_{ts}.xlsx")
    shutil.copy2(tracker_path, backup_path)
    print(f"  [OK] Backup saved: {backup_path}")
    return backup_path


def find_header_row(raw, required_col):
    for i, row in raw.iterrows():
        if any(str(v).strip() == required_col for v in row.values):
            return i
    return None


def load_sheet_with_banner(path, sheet, merge_key):
    raw = pd.read_excel(path, sheet_name=sheet, header=None)

    header_row = find_header_row(raw, merge_key)
    if header_row is None:
        raise ValueError(
            f"Could not find a '{merge_key}' column in {path} / {sheet}. "
            "Check the sheet name and column header."
        )

    raw_headers = raw.iloc[header_row].astype(str).str.strip()

    final_headers = list(raw_headers)
    if header_row > 0:
        year_row = raw.iloc[header_row - 1]
        year_ffilled = year_row.ffill()
        for idx, col_name in enumerate(raw_headers):
            if is_month_col(col_name) and not has_year_suffix(col_name):
                year_val = year_ffilled.iloc[idx] if idx < len(year_ffilled) else None
                if pd.notna(year_val) and str(year_val).strip().isdigit():
                    final_headers[idx] = f"{col_name} {str(year_val).strip()}"

    df = raw.iloc[header_row + 1:].reset_index(drop=True)
    df.columns = final_headers
    df.columns.name = None
    return df, header_row


def load_sheet_flat(path, sheet, required_col):
    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    header_row = find_header_row(raw, required_col)
    if header_row is None:
        raise ValueError(
            f"Could not find a '{required_col}' column in {path} / {sheet}. "
            "Check the sheet name and column header."
        )
    headers = raw.iloc[header_row].astype(str).str.strip()
    df = raw.iloc[header_row + 1:].reset_index(drop=True)
    df.columns = list(headers)
    df.columns.name = None
    return df, header_row


def write_sheet_in_place(tracker_path, tracker_sheet, header_row,
                          result_df, month_format=None,
                          new_month_cols=None, is_month_col_fn=None):
    wb = openpyxl.load_workbook(tracker_path)
    ws = wb[tracker_sheet]

    header_excel_row = header_row + 1
    first_data_row    = header_excel_row + 1
    n_cols            = len(result_df.columns)

    if new_month_cols and month_format == "banner" and header_excel_row > 1 and is_month_col_fn:
        year_banner_row = header_excel_row - 1
        col_year = {}
        for col_idx, col_name in enumerate(result_df.columns, start=1):
            if is_month_col_fn(col_name):
                parts = str(col_name).split()
                if len(parts) > 1 and parts[-1].isdigit():
                    col_year[col_idx] = int(parts[-1])

        if col_year:
            existing_year_merges = [
                rng for rng in list(ws.merged_cells.ranges)
                if rng.min_row == year_banner_row and rng.max_row == year_banner_row
            ]
            for rng in existing_year_merges:
                year_value = ws.cell(row=year_banner_row, column=rng.min_col).value
                if not isinstance(year_value, (int, float)):
                    continue
                year_value = int(year_value)
                matching_cols = [c for c, y in col_year.items() if y == year_value]
                if not matching_cols:
                    continue
                new_min, new_max = min(matching_cols), max(matching_cols)
                if (new_min, new_max) != (rng.min_col, rng.max_col):
                    ws.unmerge_cells(start_row=year_banner_row, start_column=rng.min_col,
                                      end_row=year_banner_row, end_column=rng.max_col)
                    if new_min != rng.min_col:
                        old_val = ws.cell(row=year_banner_row, column=rng.min_col).value
                        ws.cell(row=year_banner_row, column=rng.min_col).value = None
                        ws.cell(row=year_banner_row, column=new_min).value = old_val
                    ws.merge_cells(start_row=year_banner_row, start_column=new_min,
                                    end_row=year_banner_row, end_column=new_max)
                    print(f"  [merge] Adjusted {year_value} year-banner merge to columns {new_min}-{new_max}")

    if ws.max_row >= first_data_row:
        for row in ws.iter_rows(min_row=first_data_row, max_row=ws.max_row,
                                 min_col=1, max_col=max(ws.max_column, n_cols)):
            for cell in row:
                cell.value = None

    for col_idx, col_name in enumerate(result_df.columns, start=1):
        header_to_write = col_name
        if month_format == "banner" and is_month_col_fn and is_month_col_fn(col_name) and has_year_suffix(col_name):
            header_to_write = " ".join(str(col_name).split()[:-1])
        ws.cell(row=header_excel_row, column=col_idx, value=header_to_write)

    for row_idx, (_, data_row) in enumerate(result_df.iterrows(), start=first_data_row):
        for col_idx, col_name in enumerate(result_df.columns, start=1):
            val = data_row[col_name]
            ws.cell(row=row_idx, column=col_idx, value=None if pd.isna(val) else val)

    wb.save(tracker_path)

# ─── JOB: monthly_columns ──────────────────────────────────────────────────

def run_monthly_columns_job(job, tracker_path):
    key_cols   = job["key_cols"]
    merge_key  = job["merge_key"]
    month_fmt  = job["month_format"]
    export_path = EXPORT_PATHS[job["export_path_key"]]

    if not os.path.exists(export_path):
        raise FileNotFoundError(f"Export file not found:\n  {export_path}")

    print(f"Loading export ({job['name']})...")
    if month_fmt == "banner":
        export_df, _ = load_sheet_with_banner(export_path, job["export_sheet"], merge_key)
    else:
        export_df, _ = load_sheet_flat(export_path, job["export_sheet"], merge_key)

    print(f"Loading tracker sheet '{job['tracker_sheet']}'...")
    if month_fmt == "banner":
        tracker_df, tracker_header_row = load_sheet_with_banner(tracker_path, job["tracker_sheet"], merge_key)
    else:
        tracker_df, tracker_header_row = load_sheet_flat(tracker_path, job["tracker_sheet"], merge_key)

    if month_fmt == "banner":
        export_month_cols_raw = [c for c in export_df.columns if is_month_col(c) and not has_year_suffix(c)]
        if export_month_cols_raw:
            year_map = infer_export_month_years(export_month_cols_raw)
            export_df = export_df.rename(columns={
                c: f"{c} {year_map[c]}" for c in export_month_cols_raw if c in year_map
            })
        month_test = is_month_col
    else:
        month_test = lambda c: parse_mon_yy(c) is not None

    site_cols = [c for c in export_df.columns if c not in key_cols and str(c).strip()]
    month_cols_in_export = [c for c in site_cols if month_test(c)]
    static_cols_in_export = [c for c in site_cols if not month_test(c)]

    tracker_month_cols = [c for c in tracker_df.columns if month_test(c)]
    months_to_preserve = [c for c in tracker_month_cols if c not in month_cols_in_export]
    if months_to_preserve:
        print(f"  Preserving {len(months_to_preserve)} month column(s) no longer in export: "
              f"{', '.join(months_to_preserve)}")

    new_month_cols = [c for c in month_cols_in_export if c not in tracker_df.columns]
    if new_month_cols:
        print(f"  Adding new month column(s): {', '.join(new_month_cols)}")

    protected_cols_for_job = [c for c in PROTECTED_COLS if c in tracker_df.columns]

    for col in key_cols:
        if col not in export_df.columns:
            raise ValueError(f"Export ({job['name']}) is missing expected column '{col}'")
        if col not in tracker_df.columns:
            raise ValueError(f"Tracker sheet '{job['tracker_sheet']}' is missing expected column '{col}'")

    def make_key(row):
        return tuple(str(row[c]).strip() for c in key_cols)

    tracker_df["_key"] = tracker_df.apply(make_key, axis=1)

    dup_mask = tracker_df["_key"].duplicated(keep=False)
    duplicate_keys = sorted(set(tracker_df.loc[dup_mask, "_key"]))
    if duplicate_keys:
        print(f"  WARNING: Found exact duplicate ({' + '.join(key_cols)}) rows in tracker — "
              f"keeping the FIRST occurrence of each, extras left untouched at the bottom:")
        for k in duplicate_keys:
            print(f"       - {' / '.join(k)}")

    first_occurrence = ~tracker_df["_key"].duplicated(keep="first")
    tracker_subset = tracker_df[first_occurrence]
    if len(key_cols) == 1:
        tracker_lookup = tracker_subset.set_index(
            pd.Index([k[0] for k in tracker_subset["_key"]], name=key_cols[0])
        ).drop(columns=["_key"])
    else:
        tracker_lookup = tracker_subset.set_index(
            pd.MultiIndex.from_tuples(tracker_subset["_key"], names=key_cols)
        ).drop(columns=["_key"])

    leftover_dupe_rows = tracker_df[~first_occurrence].drop(columns=["_key"])

    updated_rows = []
    new_entries  = []
    matched_keys = set()

    for _, exp_row in export_df.iterrows():
        primary = str(exp_row[merge_key]).strip()
        if not primary or primary.lower() == "nan":
            continue

        key_tuple = make_key(exp_row)
        lookup_key = key_tuple[0] if len(key_cols) == 1 else key_tuple

        if lookup_key in tracker_lookup.index:
            row = tracker_lookup.loc[lookup_key].copy()
            for col in site_cols:
                if col in exp_row.index:
                    row[col] = exp_row[col]
            updated_rows.append(row)
            matched_keys.add(lookup_key)
        else:
            new_row = exp_row.copy()
            for col in protected_cols_for_job:
                new_row[col] = ""
            for col in months_to_preserve:
                new_row[col] = ""
            updated_rows.append(new_row)
            new_entries.append(" / ".join(key_tuple))

    untouched_keys = [k for k in tracker_lookup.index if k not in matched_keys]
    for k in untouched_keys:
        updated_rows.append(tracker_lookup.loc[k])

    result_df = pd.DataFrame(updated_rows).reset_index(drop=True)
    if not leftover_dupe_rows.empty:
        result_df = pd.concat([result_df, leftover_dupe_rows], ignore_index=True)

    original_tracker_order = [c for c in tracker_df.columns if c != "_key"]
    all_cols = list(result_df.columns)

    if new_month_cols:
        existing_months_sorted = sorted(tracker_month_cols, key=lambda c: month_sort_key(c, month_fmt))
        new_months_sorted = sorted(new_month_cols, key=lambda c: month_sort_key(c, month_fmt))

        if existing_months_sorted:
            insert_after = existing_months_sorted[0]
            insert_at = original_tracker_order.index(insert_after)
        else:
            anchor_col = next((c for c in ["Total Count", "Total Baskets"] if c in original_tracker_order), None)
            insert_at = (original_tracker_order.index(anchor_col) + 1) if anchor_col else len(original_tracker_order)

        ordered = (
            original_tracker_order[:insert_at]
            + new_months_sorted
            + original_tracker_order[insert_at:]
        )
    else:
        ordered = original_tracker_order

    ordered += [c for c in all_cols if c not in ordered]
    ordered = list(dict.fromkeys(ordered))
    result_df = result_df[[c for c in ordered if c in result_df.columns]]

    print(f"Writing '{job['tracker_sheet']}'...")
    write_sheet_in_place(
        tracker_path, job["tracker_sheet"], tracker_header_row, result_df,
        month_format=month_fmt, new_month_cols=new_month_cols, is_month_col_fn=month_test,
    )

    return {
        "name": job["name"],
        "updated": len(matched_keys),
        "untouched": len(untouched_keys),
        "new": new_entries,
        "duplicates": [" / ".join(k) for k in duplicate_keys],
        "new_months": new_month_cols,
        "preserved_months": months_to_preserve,
    }

# ─── JOB: daily_rows ──────────────────────────────────────────────────────

def run_daily_rows_job(job, tracker_path):
    key_col = job["key_cols"][0]
    export_path = EXPORT_PATHS[job["export_path_key"]]
    total_label = job["total_row_label"]
    sum_cols = job["sum_cols"]
    avg_cols = job["avg_cols"]

    if not os.path.exists(export_path):
        raise FileNotFoundError(f"Export file not found:\n  {export_path}")

    print(f"Loading export ({job['name']})...")
    export_df, _ = load_sheet_flat(export_path, job["export_sheet"], key_col)

    print(f"Loading tracker sheet '{job['tracker_sheet']}'...")
    tracker_df, tracker_header_row = load_sheet_flat(tracker_path, job["tracker_sheet"], key_col)

    is_total_row = tracker_df[key_col].astype(str).str.strip().str.lower() == total_label.lower()
    tracker_data_only = tracker_df[~is_total_row].copy()
    export_is_total = export_df[key_col].astype(str).str.strip().str.lower() == total_label.lower()
    export_data_only = export_df[~export_is_total].copy()

    data_cols = [c for c in export_data_only.columns if c not in ("No.", key_col) and str(c).strip()]

    def norm_date(v):
        if pd.isna(v):
            return ""
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        return str(v).strip()

    tracker_data_only["_key"] = tracker_data_only[key_col].apply(norm_date)
    export_data_only["_key"] = export_data_only[key_col].apply(norm_date)

    dup_mask = tracker_data_only["_key"].duplicated(keep=False)
    duplicate_keys = sorted(set(tracker_data_only.loc[dup_mask, "_key"]) - {""})
    if duplicate_keys:
        print(f"  WARNING: Found duplicate {key_col} rows in tracker — keeping the FIRST occurrence of each:")
        for k in duplicate_keys:
            print(f"       - {k}")

    first_occurrence = ~tracker_data_only["_key"].duplicated(keep="first")
    tracker_lookup = tracker_data_only[first_occurrence].set_index("_key")
    leftover_dupe_rows = tracker_data_only[~first_occurrence].drop(columns=["_key"])

    updated_rows = []
    new_dates = []
    matched_keys = set()

    for _, exp_row in export_data_only.iterrows():
        k = exp_row["_key"]
        if not k:
            continue
        if k in tracker_lookup.index:
            row = tracker_lookup.loc[k].copy()
            for col in data_cols:
                if col in exp_row.index:
                    row[col] = exp_row[col]
            updated_rows.append(row)
            matched_keys.add(k)
        else:
            updated_rows.append(exp_row.copy())
            new_dates.append(k)

    untouched_keys = [k for k in tracker_lookup.index if k not in matched_keys]

    result_df = pd.DataFrame(updated_rows)
    if "_key" in result_df.columns:
        result_df = result_df.drop(columns=["_key"])
    if not leftover_dupe_rows.empty:
        result_df = pd.concat([result_df, leftover_dupe_rows], ignore_index=True)

    result_df["_sort"] = pd.to_datetime(result_df[key_col], errors="coerce")
    result_df = result_df.sort_values("_sort", ascending=False, na_position="last").drop(columns=["_sort"])
    result_df = result_df.reset_index(drop=True)

    if "No." in result_df.columns:
        result_df["No."] = range(1, len(result_df) + 1)

    total_row = {c: None for c in result_df.columns}
    total_row[key_col] = total_label
    if "No." in result_df.columns:
        total_row["No."] = len(result_df) + 1
    for col in sum_cols:
        if col in result_df.columns:
            total_row[col] = pd.to_numeric(result_df[col], errors="coerce").sum()
    for col in avg_cols:
        if col in result_df.columns:
            total_row[col] = pd.to_numeric(result_df[col], errors="coerce").mean()

    result_df = pd.concat([result_df, pd.DataFrame([total_row])], ignore_index=True)

    original_order = [c for c in tracker_df.columns if c != "_key"]
    ordered = original_order + [c for c in result_df.columns if c not in original_order]
    ordered = list(dict.fromkeys(ordered))
    result_df = result_df[[c for c in ordered if c in result_df.columns]]

    print(f"Writing '{job['tracker_sheet']}'...")
    write_sheet_in_place(tracker_path, job["tracker_sheet"], tracker_header_row, result_df)

    return {
        "name": job["name"],
        "updated": len(matched_keys),
        "untouched": len(untouched_keys),
        "new": new_dates,
        "duplicates": duplicate_keys,
        "new_months": [],
        "preserved_months": [],
    }

# ─── MAIN ──────────────────────────────────────────────────────────────────

def main():
    print("Starting ETF Tracker Sync...")

    if not os.path.exists(TRACKER_PATH):
        raise FileNotFoundError(f"Tracker file not found:\n  {TRACKER_PATH}")

    print("Backing up tracker...")
    backup_tracker(TRACKER_PATH)

    results = []
    for job in JOBS:
        print(f"\n-- {job['name']} --")
        if job["kind"] == "monthly_columns":
            result = run_monthly_columns_job(job, TRACKER_PATH)
        elif job["kind"] == "daily_rows":
            result = run_daily_rows_job(job, TRACKER_PATH)
        else:
            raise ValueError(f"Unknown job kind: {job['kind']}")
        results.append(result)

    print(f"\nSync complete!\n")
    for r in results:
        print(f"-- {r['name']} --")
        print(f"   {r['updated']} existing rows updated")
        print(f"   {r['untouched']} rows left untouched / dropped (see notes above)")
        if r["new_months"]:
            print(f"   {len(r['new_months'])} new month column(s) added: {', '.join(r['new_months'])}")
        if r["preserved_months"]:
            print(f"   {len(r['preserved_months'])} older month column(s) preserved: {', '.join(r['preserved_months'])}")
        if r["new"]:
            print(f"   {len(r['new'])} new row(s) added:")
            for c in r["new"]:
                print(f"       - {c}")
        else:
            print("   No new rows detected")
        if r["duplicates"]:
            print(f"   WARNING: {len(r['duplicates'])} exact duplicate row(s) detected — please clean up when convenient")
        print()

    return results


if __name__ == "__main__":
    try:
        results = main()

        parts = []
        for r in results:
            parts.append(f"{r['name']}: {r['updated']} updated"
                         + (f", {len(r['new'])} new" if r["new"] else "")
                         + (f", {len(r['duplicates'])} dup(s)" if r["duplicates"] else ""))
        status_msg = " | ".join(parts)

        log_line(f"SUCCESS — {status_msg}")
        notify("ETF Tracker Sync - Success", status_msg)

    except Exception as e:
        error_msg = f"{type(e).__name__}: {e}"
        log_line(f"FAILED — {error_msg}")
        log_line(traceback.format_exc())
        notify("ETF Tracker Sync - FAILED", f"{error_msg}\nCheck {LOG_PATH}")
        print(f"\nError: {e}")
        sys.exit(1)
